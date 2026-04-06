import pandas as pd
import os
import json
from datetime import datetime
from xlsxwriter.utility import xl_col_to_name
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, NamedStyle
from openpyxl import load_workbook
from openpyxl.comments import Comment
from copy import copy
import pyambit.datamodel as mb
import re as _re


METADATA_PARAMETERS = "METADATA_PARAMETERS"
METADATA_SAMPLE_PREP = "METADATA_SAMPLE_PREP"
METADATA_SAMPLE_INFO = "METADATA_SAMPLE_INFO"


def iom_format(df, param_name="param_name", param_group="param_group"):
    df.fillna(" ", inplace=True)
    #print(df.columns)
    # Create a new DataFrame with one column
    result_df = pd.DataFrame(columns=['param_name'])
    result_df["type"] = "group"
    result_df["position"] = -1
    # Iterate through unique groups
    for group in df[param_group].unique():
        group_df = df[df[param_group] == group]
        # Get names for the group
        names = group_df[param_name].tolist()
        # Append group and names to the result DataFrame
        tmp = pd.DataFrame({'param_name': [group] + names })
        tmp["type"] = "names"
        tmp.at[0, "type"] = "group"
        tmp['position'] = -1
        result_df = pd.concat([result_df, tmp], ignore_index=True)
    return result_df


def json2frame(json_data, sortby=None):
    tmp = pd.DataFrame(json_data)
    if sortby is None:
        return tmp
    else:
        return tmp.sort_values(by=sortby)


def get_method_metadata(bp_json):
    _header = {
        "Project Work Package": bp_json.get("provenance_workpackage", ""),
        "Partner conducting test/assay": bp_json.get("provenance_provider", ""),
        "Test facility - Laboratory name": bp_json.get("provenance_provider", ""),
        "Lead Scientist & contact for test": bp_json.get("provenance_contact", ""),
        "Assay/Test work conducted by": bp_json.get("provenance_operator", ""),
        "Full name of test/assay": bp_json.get("METHOD", ""),
        "Short name or acronym for test/assay": bp_json.get("METHOD", ""),
        "Type or class of experimental test as used here": bp_json.get(
            "PROTOCOL_CATEGORY_CODE", ""),
        "End-Point being investigated/assessed by the test":  [
            item["result_name"] if "result_name" in item else "result_name_not_specified" for item in bp_json.get("question3", [])
            ],
        "End-Point units":  [item["result_unit"] if "result_unit" in item else "" for item in bp_json.get("question3", [])],
        "Raw data metrics": [item["raw_endpoint"] if "raw_endpoint" in item else "raw_endpoint_not_specified" for item in bp_json.get("raw_data_report",[])],
        "Raw data units": [item.get("raw_unit", "") for item in bp_json.get(
            "raw_data_report", [])],
        "SOP(s) for test": bp_json.get("EXPERIMENT", ""),
        "Path/link to sop/protocol": bp_json.get("EXPERIMENT_PROTOCOL", ""),
        "Test start date": bp_json.get("provenance_startdate", datetime.now()),
        "Test end date": bp_json.get("provenance_enddate", datetime.now()),
        }
    return _header


def get_materials_metadata(json_blueprint):
    sample_group_dict = {}
    for item in json_blueprint.get("METADATA_SAMPLE_INFO"):
        group = item.get("param_sample_group","DEFAULT")
        name = item["param_sample_name"]
        sample_group_dict.setdefault(group, []).append(name)    
    _header = {
        "Select item from Project Materials list": 
        sample_group_dict.get("ID", ["ID"])[0],
        "Material Name": sample_group_dict.get("NAME", ["NAME"])[0],
        "Core chemistry": sample_group_dict.get("CHEMISTRY", ["CHEMISTRY"])[0],
        "CAS No": sample_group_dict.get("CASRN", ["CAS_RN"])[0],
        "Material Supplier": 
        sample_group_dict.get("SUPPLIER", ["SUPPLIER"])[0],
        "Material State": "",
        "Batch": sample_group_dict.get("BATCH", ["BATCH"])[0],
        "Date of preparation": datetime.now()
    }
    return _header


def get_materials_columns(nanomaterial=True):
    if nanomaterial:
        return ["", "ERM identifier", "ID", "Name", "CAS", "type", "Supplier",
                "Supplier code", "Batch", "Core", "BET surface in m²/g"]
    else:
        return ["", "Material identifier", "ID", "Name", "CAS", "type",
                "Supplier", "Supplier code", "Batch", "Core"]


def get_treatment(json_blueprint):
    _maxfields = 15
    tmp = []
    condition_type = None
    for item in json_blueprint.get("conditions",[]):
        name = "conditon_name"
        isreplicate = item["condition_type"].startswith("c_replicate")
        isconcentration = item["condition_type"].startswith("c_concentration")
        if not isreplicate:
            tmp.append({'param_name': "TREATMENT {}".format(item[name].upper()),
                         'type': 'group', 'position': '0', 'position_label': 0,
                         'datamodel': item['condition_type'], "value": ""})
        else:
            if condition_type != isreplicate:
                tmp.append({'param_name': "CONTROLS", 'type': 'group',
                            'position': '0', 'position_label': 0,
                            'datamodel': "c_replicate",
                            "value" : ""})
                tmp.append({'param_name': "Positive controls abbreviations", 
                            'type': 'names', 'position': '0', 'position_label': 0, 
                            'datamodel': "CONTROL", "value": ""})
                tmp.append({'param_name': "Positive controls description",
                            'type': 'names', 'position': '0', 'position_label': 0,
                            'datamodel': "CONTROL", "value": ""})
                tmp.append({'param_name': "Negative controls abbreviations",
                            'type': 'names', 'position': '0', 'position_label': 0,
                            'datamodel': "CONTROL", "value": ""})
                tmp.append({'param_name': "Negative controls description", 
                            'type': 'names', 'position': '0',
                            'position_label': 0, 'datamodel': "CONTROL",
                             "value": ""})
                tmp.append({'param_name': "REPLICATES", 'type': 'group',
                            'position': '0', 'position_label': 0,
                            'datamodel': "c_replicate", "value": ""})
        if "condition_unit" in item:
            tmp.append({'param_name': "{} series unit".format(item[name]),
                        'type': 'names', 'position': '0', 'position_label': 0,
                        'datamodel': item['condition_type'], "value": item["condition_unit"]})
        if not isreplicate:
            tag = item['condition_type'].split('_')[1][0].upper()
            _start = 0 if isconcentration else 1
            tmp.append({'param_name': "{} series labels".format(item[name]), 
                        'type': 'names', 'position': '0', 'position_label': 0,
                        'datamodel': item['condition_type'],
                        "value": [f"{tag}{i}" if i <= 3 else "" for i in range(1, _maxfields + 1)]})
        else:
            _start = 0
        tmp.append({'param_name': "{}".format(item[name]), 'type': 'names',
                    'position': '0', 'position_label': 0, 'datamodel': item['condition_type'], 
                    "value":  [i if i<=(2+_start) else "" for i in range(_start, _maxfields + _start + 1)]})
        if isconcentration:
            tmp.append({'param_name': "Treatment type series", 'type': 'names',
                        'position': '0', 'position_label': 0,
                        'datamodel': "c_treatment", "value": ""})
        condition_type = isreplicate
    return pd.DataFrame(tmp)


def customize_config(json_blueprint, nmpconfig):

    df_info, df_result, df_raw, df_conditions, df_calibrate = get_template_frame(
        json_blueprint)
    nmpconfig.get("TEMPLATE_INFO")["NAME"] = json_blueprint.get("template_name","")
    for tag in ["template_author","template_author_orcid","template_acknowledgment",
                "template_layout","template_status","template_date"]:
        nmpconfig.get("TEMPLATE_INFO")[tag] = json_blueprint.get(tag,"")
    offset = 7
    papp = nmpconfig.get("PROTOCOL_APPLICATIONS")[0]
    papp["PROTOCOL_CATEGORY_CODE"] = json_blueprint.get("PROTOCOL_CATEGORY_CODE")
    papp["PROTOCOL_TOP_CATEGORY"] = json_blueprint.get("PROTOCOL_TOP_CATEGORY")

    params = papp.get("PARAMETERS",{})
    df_info["param_type"] = None
    for param in json_blueprint.get(METADATA_PARAMETERS,[]):
        _name = param.get("param_name",None)
        df_info.loc[df_info["param_name"] == _name, "unit"] = param.get("param_unit",None)
        df_info.loc[df_info["param_name"] == _name, "param_type"] = param.get("param_type","")
        df_info.loc[df_info["param_name"] == _name, "param_group"] = param.get("param_group","")

    for row in json_blueprint.get("question3",[]):
        _name = param.get("result_name",None)
        df_result.loc[df_result["result_name"] == _name, "result_aggregate"] = param.get("result_aggregate",None)
        df_result.loc[df_result["result_name"] == _name, "result_endpoint_uncertainty"] = param.get("result_endpoint_uncertainty",None)

    for index, row in df_info.iterrows():
        _tmp = {"ITERATION": "ABSOLUTE_LOCATION",  "SHEET_INDEX": 1, "COLUMN_INDEX": "B"}
        if row["datamodel"] in [METADATA_PARAMETERS, METADATA_SAMPLE_PREP]:
            if row["type"] == "names":
                _tmp["ROW_INDEX"] = row["position"] + offset
                params[row["param_name"]] = _tmp
        if row["param_name"] == "Full name of test/assay":
            params["Assay"]["ROW_INDEX"] = row["position"] + offset

    unique_conditions = sorted(set(condition for conditions in df_result["results_conditions"].dropna() for condition in conditions))
    # results table
    sheet_index = 1
    effects = []
    for result_type, name, unit,aggregate,errqualifier, df in [
        ("RAW_DATA","raw_endpoint","raw_unit","raw_aggregate","raw_endpoint_uncertainty", df_raw), 
        ("AGGREGATED", "result_name","result_unit", "result_aggregate","result_endpoint_uncertainty", df_result)]:
        if df is None:
            continue
        sheet_index = sheet_index + 1
        for index, row in df.iterrows():
            effect = {
                        "ENDPOINT": row[name].upper().replace(" ","_"),
                        "ENDPOINT_TYPE" : result_type,
                        "VALUE" : {
                            "ITERATION": "ROW_SINGLE",
                            "SHEET_INDEX": sheet_index,
                            "COLUMN_INDEX":  chr(65 + len(unique_conditions) + index +1)
                        },
                        "UNIT": row[unit].upper().replace(" ","_"),
                        "CONDITIONS" : {}
                    }
            if errqualifier in row and not pd.isna(row[errqualifier]):
                effect["ERR_QUALIFIER"] = row[errqualifier]
            if aggregate in row and not pd.isna(row[aggregate]):
                effect["ENDPOINT_TYPE"] = row[aggregate]

            for i, c in enumerate(unique_conditions):
                row = df_conditions.loc[df_conditions["conditon_name"] == c]
                if row.empty:
                    continue
                row = row.iloc[0]
                _tmp = { "COLUMN_INDEX":  chr(65 + i +1), "ITERATION": "ROW_SINGLE", "SHEET_INDEX": 2}
                if not pd.isna(row["condition_unit"]):
                    _tmp["UNIT"] = row["condition_unit"]
                effect["CONDITIONS"][c.upper()] = _tmp

            effects.append(effect)
    papp["EFFECTS"] = effects
    return nmpconfig


def get_nmparser_config(json_blueprint):
    current_directory = os.path.dirname(os.path.abspath(__file__))
    json_file_path = os.path.join(current_directory, 
                                  "../../resource/nmparser/DEFAULT_TABLE.json")
    config = {}
    with open(json_file_path, 'r') as json_file:
        # Load the JSON data from the file
        config = json.load(json_file)
    try:
        nmpconfig = customize_config(json_blueprint, config)
        return nmpconfig
    except Exception as err:
        print(err)


def create_nested_headers_dataframe(dicts,
                                    keys={METADATA_PARAMETERS: {'group': 'param_group', 'name': 'param_name', 'unit': 'param_unit'}},
                                    levels=['group', 'name', 'unit'],
                                    lookup={METADATA_SAMPLE_INFO: "Sample", 
                                            METADATA_SAMPLE_PREP: "Sample preparation",
                                        "OTHER_SAMPLEPREP": "",
                                        "raw_data_report": "Raw data", "question3": "Results"},
                                    condition_field=[ "raw_conditions","results_conditions"]
                                    ):
    # Initialize an empty DataFrame
    df = pd.DataFrame()
    # Build global condition metadata lookup
    condition_meta = {
        cond.get("conditon_name"): {
            "unit": cond.get("condition_unit", ""),
            "type": cond.get("condition_type", "")
        }
        for cond in dicts.get("conditions", [])
        if cond.get("conditon_name")
    }    

    # Iterate through the dictionaries
    key_conditions = set()    
    for key in keys:
        params = dicts.get(key, [])
        # Collect all unique conditions for this key
        for param in params:
            for cf in condition_field:
                if cf in param:
                    key_conditions.update(param.get(cf, []))

    # Add one column per unique condition (once per key)
    for cond_name in key_conditions:
        cond_info = condition_meta.get(cond_name, {})
        cond_tags = ["Experimental factors"]
        cond_tags.append(cond_name)
        cond_tags.append("")
        cond_tags.append(cond_info.get("unit",""))
        df[tuple(cond_tags)] = None

    for key in keys:
        params = dicts.get(key, [])
        top_label = lookup.get(key, key)
        for param in params:
            try:
                tags = [top_label]
                for level in levels:
                    _tmp = param.get(keys[key][level], "")
                    tags.append(lookup.get(_tmp, _tmp) )
                df[tuple(tags)] = None
            except Exception as err:
                print(f"Error processing param: {e}")
                continue


    # Create MultiIndex DataFrame
    names = ['']
    names.extend(levels)
    df.columns = pd.MultiIndex.from_tuples(df.columns, names=names)
    return df


def autofit_columns(sheet, cols=None):
    # Autofit column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust for padding and scaling
        sheet.column_dimensions[column_letter].width = adjusted_width    
        # Apply colors to top-level keys
    top_level_colors = {'METADATA_PARAMETERS': 'BDD7EE',
                        'Sample': 'FCE4D6', 'Sample preparation': 'BDD7EE',
                        'Raw data': 'FCE4D6', 'Results': 'BDD7EE'}

    for col_num, value in enumerate(cols):
        top_level = value[0]
        if top_level in top_level_colors:
            clr = top_level_colors.get(top_level, "white")
            pf = PatternFill(start_color=clr, end_color=clr, fill_type="solid")
            sheet.cell(row=1, column=col_num+1).fill = pf

        
def autofit_multilevel(df, worksheet):
    for idx, col in enumerate(df.columns):
        # Find the maximum length of the column header (using the last level of multi-index)
        max_length = max(len(str(level)) for level in col) + 1
        # Set the column width based on the length of the column header
        worksheet.set_column(idx, idx, max_length)


def pchem_format_2excel(file_path_xlsx, json_blueprint):
    _SHEET_INFO = "Provider_informations"
    _SHEET_RAW = "Raw_data_TABLE"
    _SHEET_RESULT = "Results_TABLE"
    _SHEET_MATERIAL = "Materials"
    _SHEET_MEASUREMENT = "Experimental_setup"
    current_script_directory = os.path.dirname(os.path.abspath(__file__))
    #resource_file = os.path.join(current_script_directory, "../../resource/nmparser","template_pchem.xlsx")
    #shutil.copy2(resource_file, file_path_xlsx)
    with pd.ExcelWriter(file_path_xlsx, engine='xlsxwriter', mode='w') as writer:
        # sheet = writer.book["Provider_informations"]
        worksheet = writer.book.add_worksheet(_SHEET_INFO)
        bold_format = writer.book.add_format({'bold': True})
        orange_bg_format = writer.book.add_format({'bg_color': '#FFF2CC'})
        material_format = writer.book.add_format({'bg_color': '#00B0F0'})
        position_format = writer.book.add_format({'bg_color': '#FFC000'})
        #_colors = { "top" : '#002060' , "orange" : '#FFC000', 'skyblue' : '#00B0F0' , 'grey' : '#C0C0C0', 'input' : '#FFF2CC'}
        for t in [("A2", "General information"),
                  ("A10", "Template name"),("A11", "Template version "), ("A12", "Template authors"),
                  ("A13", "Template acknowledgment"), ("A6", "Project"), ("A7", "Workpackage"),
                  ("A8", "Partner"), ("D7", "Study"), ("A15", "Template downloaded"),
                  ("B15", datetime.now().strftime("%Y-%m-%d"))
                 ]:
            worksheet.write(t[0], t[1], bold_format)                
        for t in [("E7", "METHOD"),("C2", "EXPERIMENT"), ("B10", "template_name"),
                  ("B11", "template_status"), ("B12", "template_author"),
                  ("B13", "template_acknowledgment"), ("B6", "provenance_project"),
                  ("B7", "provenance_workpackage"), ("B8","provenance_provider")]:
            worksheet.write(t[0], json_blueprint.get(t[1], ""), orange_bg_format)

        worksheet = writer.book.add_worksheet(_SHEET_RESULT)
        df = create_nested_headers_dataframe(json_blueprint,
                                             keys={"raw_data_report": {
                                                 'name': 'raw_endpoint',
                                                 'type': 'raw_aggregate',
                                                 'unit': 'raw_unit'},
                                                 "question3": {
                                                     'name': 'result_name',
                                                     'type': 'result_aggregate',
                                                     'unit': 'result_unit'}},
                                             levels=['name', 'type', 'unit'],
                                             lookup={
                                                 "raw_data_report": "Raw data",
                                                 "question3": "Results"},
                                             condition_field=[ "raw_conditions","results_conditions"]
                                            )
        #df.insert(0, 'Material ID',None)
        df.insert(0, 'Position_ID',None)
        df.to_excel(writer, sheet_name=_SHEET_RESULT) 
        worksheet = writer.book.get_worksheet_by_name(_SHEET_RESULT)
        worksheet.write('A1', 'Material ID', material_format)
        worksheet.write('A2', ' ', material_format)
        worksheet.write('A3', ' ', material_format)
        worksheet.write('A4', ' ', material_format)      
        worksheet.write('B1', 'Position_ID', position_format)
        worksheet.write('B2', ' ', position_format)
        worksheet.write('B3', ' ', position_format)
        worksheet.write('B4', ' ', position_format)            
        autofit_multilevel(df, worksheet)

        df = create_nested_headers_dataframe(json_blueprint,
                                             keys={"METADATA_PARAMETERS" : {'group' : 'param_group', 'name' : 'param_name', 'unit' : 'param_unit'}})
        
        df.to_excel(writer, sheet_name=_SHEET_MEASUREMENT)
        worksheet = writer.book.get_worksheet_by_name(_SHEET_MEASUREMENT)
        worksheet.write('A1', 'Position_ID',position_format)
        worksheet.write('A2', ' ', position_format)
        worksheet.write('A3', ' ', position_format)
        worksheet.write('A4', ' ', position_format)     
        worksheet.write('A5', 'P1', position_format)     
        worksheet.write('A6', 'P2', position_format)
        position_identifiers_range = "{}!$A5:$A1048576".format(_SHEET_MEASUREMENT)  # Entire column A
        writer.book.define_name('Position_Identifiers', position_identifiers_range)           
        autofit_multilevel(df, worksheet)   
        validation = {
            'validate': 'list',
            'source': '=Position_Identifiers'
        }
        writer.book.get_worksheet_by_name(_SHEET_RESULT).data_validation("B5:B1048576", validation)        

        df = create_nested_headers_dataframe(json_blueprint, keys={
                #"METADATA_SAMPLE_INFO" : {'group' : 'param_sample_group', 'name' : 'param_sample_name'},
                "METADATA_SAMPLE_PREP" : {'group' : 'param_sampleprep_group', 'name' : 'param_sampleprep_name'}},
                levels=['group','name'], 
                lookup={'METADATA_SAMPLE_INFO' : "Sample", "METADATA_SAMPLE_PREP" : "Sample preparation", "group" : ""})
        df.to_excel(writer, sheet_name="SAMPLES")     
        worksheet = writer.book.get_worksheet_by_name("SAMPLES")
        worksheet.write('A1', 'Material ID', material_format)
        worksheet.write('A2', ' ', material_format)
        worksheet.write('A3', ' ', material_format)
        validation = {
            'validate': 'list',
            'source': '=ERM_Identifiers'
        }
        writer.book.get_worksheet_by_name("SAMPLES").data_validation("A4:A1048576", validation)          
        autofit_multilevel(df, worksheet)
        materials_sheet = create_materials_sheet(
            writer.book, writer, materials=_SHEET_MATERIAL, info=None,
            results=[_SHEET_RESULT], material_column="A5:A1048576")
    add_hidden_jsondef(file_path_xlsx, json_blueprint)


def add_hidden_jsondef(file_path_xlsx, json_blueprint):
    try:
        workbook = load_workbook(file_path_xlsx)
        hidden_sheet = workbook.create_sheet("TemplateDesigner")
        hidden_sheet.sheet_state = 'hidden'
        hidden_sheet['A1'] = "uuid"
        hidden_sheet['B1'] = "surveyjs"
        hidden_sheet['A2'] = json_blueprint.get("template_uuid", "")
        hidden_sheet['B2'] = json.dumps(json_blueprint)
        hidden_sheet['A3'] = "version"
        hidden_sheet['B3'] = "1.01"        
        hidden_sheet['B2'].style = NamedStyle(name='hidden', hidden=True)  # Hide the cell
        workbook.save(file_path_xlsx)
    except Exception as err:
        print(err)  


def add_plate_layout(file_path_xlsx, json_blueprint):
    if "data_platelayout" in json_blueprint.get("data_sheets", []):
        platexlsx = "platelayout_{}well.xlsx".format(json_blueprint.get("plate_format", 96) )
        current_script_directory = os.path.dirname(os.path.abspath(__file__))
        resource_file = os.path.join(current_script_directory, "../../resource/nmparser", platexlsx)
        copy_sheets(resource_file, file_path_xlsx)


def copy_sheets(source_file, destination_file):
    # Load the source Excel file
    source_wb = load_workbook(source_file)
    # Load the destination Excel file
    destination_wb = load_workbook(destination_file)
    
    # Iterate over each sheet in the source Excel file
    for sheet_name in source_wb.sheetnames:
        # Get the source sheet
        source_sheet = source_wb[sheet_name]
        # Create a new sheet in the destination file with the same name
        destination_sheet = destination_wb.create_sheet(sheet_name)
        # Iterate over each row in the source sheet
        for row in source_sheet.iter_rows(values_only=True):
            destination_sheet.append(row)
        # Copy formulas from the source sheet to the destination sheet
        for row in source_sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    destination_sheet[cell.coordinate].value = cell.value
                if cell.comment:
                    # Create a new comment on the destination cell
                    destination_sheet[cell.coordinate].comment = Comment(cell.comment.text, cell.comment.author)
                cell_dst = destination_sheet[cell.coordinate]    
                cell_dst.font = copy(cell.font)
                cell_dst.fill = copy(cell.fill)
                cell_dst.border = copy(cell.border)
                cell_dst.alignment = copy(cell.alignment)
                cell_dst.number_format = copy(cell.number_format)
                cell_dst.protection = copy(cell.protection)

    destination_wb.save(destination_file)


def get_template_frame(json_blueprint):
    if "METADATA_SAMPLE_INFO" in json_blueprint:
        df_sample = pd.DataFrame(list(get_materials_metadata(json_blueprint).items()), columns=['param_name', 'value'])
        #df_sample = json2frame(json_blueprint["METADATA_SAMPLE_INFO"],sortby=["param_sample_group"]).rename(columns={'param_sample_name': 'param_name'})
        df_sample["type"] = "names"
        df_sample["position"] = -1
        df_sample["datamodel"] = "METADATA_SAMPLE_INFO"
        df_sample = pd.concat([pd.DataFrame([{'param_name': "Test Material Details", 'type': 'group', 'position' : '0', 'position_label' : 0,'datamodel' : 'METADATA_SAMPLE_INFO', 'value' : ''}],
                                            columns=df_sample.columns), df_sample], ignore_index=True)
    else:
        raise Exception("Missing METADATA_SAMPLE_INFO")

    if "METADATA_SAMPLE_PREP" in json_blueprint:
        df_sample_prep = json2frame(json_blueprint["METADATA_SAMPLE_PREP"],sortby=["param_sampleprep_group"]).rename(columns={'param_sampleprep_name': 'param_name'})
        result_df_sampleprep = iom_format(df_sample_prep, "param_name", "param_sampleprep_group")
        result_df_sampleprep["datamodel"] = "METADATA_SAMPLE_PREP"
        result_df_sampleprep["value"] = ""
    else:
        raise Exception("Missing METADATA_SAMPLE_PREP")
    if "METADATA_PARAMETERS" in json_blueprint:
        df_params = json2frame(json_blueprint["METADATA_PARAMETERS"], sortby=["param_group"])
        result_df = iom_format(df_params)
        result_df["datamodel"] = "METADATA_PARAMETERS"
        result_df["value"] = ""
    else:
        raise Exception("Missing METADATA_PARAMETERS")

    #print(df_sample.columns,result_df.columns)
    #empty_row = pd.DataFrame({col: [""] * len(result_df.columns) for col in result_df.columns})
    treatment = get_treatment(json_blueprint)

    df_method = pd.DataFrame(list(get_method_metadata(json_blueprint).items()), columns=['param_name', 'value'])
    df_method["type"] = "names"
    df_method["position"] = -1
    df_method["datamodel"] = "METHOD"
    for df in [df_method, df_sample, result_df_sampleprep, result_df, treatment]:
        if not ("value" in df.columns):
            print(df.columns)
    df_info = pd.concat([
        df_method[["param_name", "type", "position", "datamodel", "value"]],
        df_sample[["param_name", "type", "position", "datamodel", "value"]],
        result_df_sampleprep[["param_name", "type", "position", "datamodel", "value"]],
        result_df[["param_name", "type", "position", "datamodel", "value"]],
        treatment[["param_name", "type", "position", "datamodel", "value"]]
        ], ignore_index=True)
    #print(df_info)
#:END: Please do not add information below this line
#Template version	{{ || version }}
#Template authors	{{ || acknowledgements }}
#Template downloaded	{{ || downloaded }}
    df_info["position"] = range(1, 1 + len(df_info) )
    df_info["position_label"] = 0
    df_info = pd.concat([df_info, pd.DataFrame([{ "param_name" : "Linked exeriment identifier", "type" : "names", "position" : 1, "position_label" : 5 , "datamodel" : "INVESTIGATION_UUID","value" : ""}])])
    df_conditions = pd.DataFrame(json_blueprint["conditions"])
    if "data_sheets" not in json_blueprint:
        json_blueprint["data_sheets"] = ["data_processed"]
    if "data_processed" in json_blueprint["data_sheets"]:    
        df_result = pd.DataFrame(json_blueprint["question3"]) if 'question3' in json_blueprint else None
    else:
        df_result = None
    if "data_raw" in json_blueprint["data_sheets"]:
        df_raw = pd.DataFrame(json_blueprint["raw_data_report"]) if "raw_data_report" in json_blueprint else None
    else:
        df_raw = None
    if "data_calibration" in json_blueprint["data_sheets"]:
        df_calibration = pd.DataFrame(json_blueprint["calibration_report"]) if "calibration_report" in json_blueprint else None
    else:
        df_calibration = None    
    return df_info, df_result, df_raw, df_conditions, df_calibration


def get_unit_by_condition_name(json_blueprint, name):
    for condition in json_blueprint['conditions']:
        if condition['condition_name'] == name:
            return condition.get('condition_unit', None)
    return None


def results_table(df_result, df_conditions=None,
                  result_name='result_name',
                  result_unit='result_unit',
                  results_conditions='results_conditions', sample_column="Material"):
    result_names = df_result[result_name]
    try:
        result_unit = df_result[result_unit]
    except Exception as err:
        result_unit = None

    header1 = list([sample_column])
    header2 = list([""])
    if results_conditions in df_result.columns:
        unique_conditions = sorted(set(condition for conditions in df_result[results_conditions].dropna() for condition in conditions))
        header1 = header1 + list(unique_conditions)
        for c in list(unique_conditions):
            try:
                unit = df_conditions.loc[df_conditions['conditon_name'] == c, 'condition_unit'].iloc[0]
                header2 = header2 + [unit if not pd.isnull(unit) else ""]
            except Exception:
                header2 = header2 + [""]

    header1 = header1 + list(result_names)
    if result_unit is not None:
        header2 = header2 + list(result_unit)
        return pd.DataFrame([header2], columns=header1)
    else:
        return pd.DataFrame(columns=header1)


def iom_format_2excel(
        file_path, df_info, df_result, 
        df_raw=None, df_conditions=None, df_calibration=None):
    _SHEET_INFO = "Test_conditions"
    _SHEET_RAW = "Raw_data_TABLE"
    _SHEET_RESULT = "Results_TABLE"
    _SHEET_CALIBRATION = "Calibration_TABLE"
    _SHEET_MATERIAL = "Materials"
    _guide = [
    "Please complete all applicable fields below as far as possible. Aim to familiarise yourself with the Introductory Guidance and Example Filled Templates.",
    "While aiming to standardise data recording as far as we can, flexibility may still be needed for some Test/Assay types and their results:",
    "Thus it may be necessary to add additional items e.g. for further replicates, concentrations, timepoints, or other variations on inputs, results outputs, etc.",
    "If so, please highlight changes & alterations e.g. using colour, and/or comments in notes, or adjacent to data/tables to flag items, fluctuations from norm, etc."
    ]
    _colors = { "top": '#002060' , "orange": '#FFC000', 'skyblue': '#00B0F0' , 'grey': '#C0C0C0', 'input': '#FFF2CC'}
    with pd.ExcelWriter(file_path, engine='xlsxwriter', mode='w') as writer:
        startrow = 7
        _sheet = _SHEET_INFO
        workbook = writer.book
        worksheet = workbook.add_worksheet(_sheet)
        worksheet.set_column(1, 1, 20)
        #writer.sheets[_sheet]
        cell_format_def = {
                    "group":  {'bg_color': _colors['grey'], 'font_color': 'blue', 'text_wrap': True, 'bold': True},
                    "names": {'bg_color': _colors['input'], 'text_wrap': True, 'align': 'right'},
                    "group_labels": {'bg_color': _colors['grey'], 'font_color': 'blue', 'text_wrap': True, 'bold': True},
                    "names_labels": { 'align': 'right', 'bold': True},
                    "top1": {'bg_color': _colors["top"], 'font_color': 'white', 'text_wrap': False, 'font_size': 14, 'bold': True},
                    "top7": {'bg_color': _colors["top"], 'font_color': 'white', 'text_wrap': False, 'font_size': 11, 'bold': True},
                    "orange": {'bg_color': _colors["orange"], 'font_color': 'blue', 'text_wrap': False, 'font_size': 12, 'bold': True},
                    "skyblue": {'bg_color': _colors["skyblue"], 'text_wrap': False}
                    }
        cell_format = {}
        for cf in cell_format_def:
            cell_format[cf] = workbook.add_format(cell_format_def[cf])

        for p in df_info['position_label'].unique():
            max_length = df_info.loc[df_info["position_label"] == p]["param_name"].apply(lambda x: len(str(x))).max()
            worksheet.set_column(p, p, max_length + 1)
            worksheet.set_column(p+1, p+1, 20)

        for index, row in df_info.iterrows():
            cf = cell_format[row["type"]]
            cf_labels = cell_format["{}_labels".format(row["type"])]
            worksheet.write(startrow+row['position']-1, row['position_label'], row['param_name'], cf_labels)
            if isinstance(row["value"], datetime):
                vals = [row["value"].strftime("%Y-%m-%d")]
            else:
                vals = row["value"] if isinstance(row["value"], list) else [str(row["value"])]
            for index, value in enumerate(vals):
                worksheet.write(startrow+row['position']-1, row['position_label']+index+1, value, cf)
            if row["type"] == "group":
                worksheet.set_row(startrow+row['position']-1, None, cf_labels)
            else:
                try:
                    worksheet.write_comment(startrow+row['position']-1, row['position_label']+1, row["datamodel"])
                except Exception:
                    #print(row['param_name'],row["datamodel"])
                    pass

        for row in range(1, startrow-2):
            worksheet.set_row(row, None, cell_format["top7"])
            worksheet.write(row, 0, _guide[row-1])

        worksheet.set_row(startrow-2, None, cell_format["orange"])
        worksheet.set_row(startrow-1, None, cell_format["skyblue"])
        worksheet.write("A1", "Project")
        worksheet.write("B1", "Test Data Recording Form (TDRF)")
        worksheet.write("A6", "TEST CONDITIONS")
        worksheet.write("B6", "Please ensure you also complete a Test Method Description Form (TMDF) for this test type")

        #conc_range = "{}!$B$72:$G$72".format(_SHEET_INFO)  # Entire column B
        #workbook.define_name('CONCENTRATIONS', conc_range)
        linksheets = []
        if df_raw is not None:
            _sheet = _SHEET_RAW
            linksheets = [_sheet]
            new_df = results_table(df_raw, df_conditions,
                                    result_name='raw_endpoint',
                                    result_unit= 'raw_unit',
                                    results_conditions='raw_conditions')
            new_df.to_excel(writer, sheet_name=_sheet, index=False, freeze_panes=(2, 0))
            worksheet = writer.sheets[_sheet]
            #print(new_df.columns)
            for i, col in enumerate(new_df.columns):
                worksheet.set_column(i, i, len(col) + 1 )
                if col == "concentration":
                    colname = xl_col_to_name(i)
                    #worksheet.data_validation('{}3:{}1048576'.format(colname,colname), 
                    #                          {'validate': 'list',
                    #                      'source': '=CONCENTRATIONS'})
    
            #worksheet.add_table(3, 1, 1048576, len(new_df.columns), {'header_row': True, 'name': _SHEET_RAW})

        if df_result is not None:
            _sheet = _SHEET_RESULT 
            new_df = results_table(df_result, result_name='result_name', 
                                   results_conditions='results_conditions')
            new_df.to_excel(writer, sheet_name=_sheet, index=False, 
                            freeze_panes=(2, 0))
            worksheet = writer.sheets[_sheet]
            for i, col in enumerate(new_df.columns):
                worksheet.set_column(i, i, len(col) + 1 )
            linksheets.append(_sheet)

        if df_calibration is not None:
            _sheet = _SHEET_CALIBRATION 
            new_df = results_table(df_calibration, result_name='calibration_entry', 
                                   result_unit="calibration_unit",
                                   results_conditions='calibration_conditions',
                                   sample_column="Sample")
            new_df.to_excel(writer, sheet_name=_sheet, index=False, 
                            freeze_panes=(2, 0))
            worksheet = writer.sheets[_sheet]
            for i, col in enumerate(new_df.columns):
                worksheet.set_column(i, i, len(col) + 1 )
            linksheets.append(_sheet)            

        materials_sheet = create_materials_sheet(
            workbook, writer, materials=_SHEET_MATERIAL,
            info=_SHEET_INFO, results=linksheets)
        #debug
        #df_info.to_excel(writer, sheet_name="df_info", index=False, freeze_panes=(2, 0))
        #if df_result is not None:
        #    df_result.to_excel(writer, sheet_name="df_result", index=False, freeze_panes=(2, 0))
        #if df_raw is not None:
        #    df_raw.to_excel(writer, sheet_name="df_raw", index=False, freeze_panes=(2, 0))
        #df_conditions.to_excel(writer, sheet_name="df_conditions", index=False, freeze_panes=(2, 0))


def create_materials_sheet(workbook, writer, materials, info=None, results=[], material_column="A3:A1048576"):
    info_sheet = None if info is None else writer.sheets[info]
    materials_sheet = workbook.add_worksheet(materials)
    column_headers = get_materials_columns()
    table = pd.DataFrame(columns=column_headers)
    table.to_excel(writer, sheet_name=materials, startrow=0, startcol=0, index=False)
    erm_identifiers_range = "{}!$B:$B".format(materials)  # Entire column B
    workbook.define_name('ERM_Identifiers', erm_identifiers_range)
    validation_cell = 'B25'  # cell to apply validation
    validation = {
        'validate': 'list',
        'source': '=ERM_Identifiers'
    }
    if info_sheet is not None:
        info_sheet.data_validation(validation_cell, validation)
        vlookup = [('B26',3),('B27',9),('B28',4),('B29',6),('B31',8)]
        for v in vlookup:
            formula = '=VLOOKUP(B$25,Materials!$B:$J,"{}",FALSE)'.format(v[1])
            info_sheet.write_formula(v[0], formula)
        readonly_format = workbook.add_format({'locked': True})
    for result in results:
        try:
            result_sheet = writer.sheets[result]
            result_sheet.data_validation(material_column, validation)
            #protect_headers(result_sheet,readonly_format)
        except Exception as err:
            print(err)
            pass
    return materials_sheet


def protect_headers(worksheet, readonly_format):
    worksheet.set_default_row(options={'locked': False})
    worksheet.set_row(0, None, readonly_format)
    worksheet.set_row(1, None, readonly_format)
    worksheet.protect()


def get_datamodel(json_blueprint):
    df_info, df_result, df_raw, df_conditions = get_template_frame(
        json_blueprint)
    params = {}
    for index, row in df_info.iterrows():
        if row["datamodel"] == METADATA_PARAMETERS:
            #unit = row["condition_unit"]
            params[row["param_name"]] = mb.Value()
    conditions = {} 
    for index, row in df_conditions.iterrows():
        unit = row["condition_unit"]
        conditions[row["condition_name"]] = mb.Value(unit=unit)

    effects = []
    for index, row in df_raw.iterrows():
        effects.append(mb.EffectRecord(
            endpoint=row["result_name"],
            endpointtype="RAW_DATA",
            result=mb.EffectResult(
                unit=row["result_unit"]
            ),
            conditions=conditions,
            sampleID="sample123",
        ))
    for index, row in df_result.iterrows():
        effects.append(mb.EffectRecord(
            endpoint=row["result_name"],
            endpointtype="type",
            result=mb.EffectResult(
                unit=row["result_unit"]
            ),
            conditions=conditions,
            sampleID="sample123",
        ))    

    protocol = mb.Protocol(
        topcategory="TOX",
        category=mb.EndpointCategory(code="ABC123"),
        endpoint="Some endpoint",
        guideline=["Rule 1", "Rule 2"],
    )
    citation = mb.Citation(year=2024, title="Sample Title", owner="Sample Owner")    
    papp = mb.ProtocolApplication(
        uuid="123e4567-e89b-12d3-a456-426614174000",
        interpretationResult="Result",
        interpretationCriteria="Criteria",
        parameters=params,
        citation=citation,
        effects=effects,
        owner=mb.SampleLink.create(
            sample_uuid="sample-uuid", sample_provider="Sample Provider"
        ),
        protocol=protocol,
        investigation_uuid="investigation-uuid",
        assay_uuid="assay-uuid",
        updated="2024-08-15",
    )
    return papp

def get_parameters(json_blueprint):
    
    return None


def apply_blueprint_customizations(df_info, df_result, df_conditions, json_blueprint):
    """
    Fill df_info, df_result, and df_conditions with default/custom values
    from the blueprint before writing to Excel.
    """
    # --- Method metadata ---
    method_meta = get_method_metadata(json_blueprint)
    df_info.loc[df_info['datamodel'] == 'METHOD', 'value'] = df_info.loc[df_info['datamodel'] == 'METHOD', 'param_name'].map(
        lambda x: method_meta.get(x, "")
    )

    # --- Sample metadata ---
    sample_meta = get_materials_metadata(json_blueprint)
    df_info.loc[df_info['datamodel'] == METADATA_SAMPLE_INFO, 'value'] = df_info.loc[df_info['datamodel'] == METADATA_SAMPLE_INFO, 'param_name'].map(
        lambda x: sample_meta.get(x, "")
    )

    # --- Sample prep metadata ---
    for prep in json_blueprint.get(METADATA_SAMPLE_PREP, []):
        mask = df_info['param_name'] == prep.get('param_sampleprep_name')
        df_info.loc[mask, 'value'] = prep.get('default_value', "")

    # --- Parameters ---
    for param in json_blueprint.get(METADATA_PARAMETERS, []):
        mask = df_info['param_name'] == param.get('param_name')
        df_info.loc[mask, 'value'] = param.get('default_value', "")

    # --- Treatments ---
    treatment_df = get_treatment(json_blueprint)
    for idx, row in df_info.iterrows():
        if row['param_name'] in treatment_df['param_name'].values:
            df_info.at[idx, 'value'] = treatment_df.loc[treatment_df['param_name'] == row['param_name'], 'value'].values[0]

    # --- Pre-fill results table if defaults exist ---
    if df_result is not None:
        for res in json_blueprint.get("question3", []):
            mask = df_result['result_name'] == res['result_name']
            if mask.any() and 'default_value' in res:
                # Fill first condition column with default
                conditions = res.get('results_conditions', [])
                if conditions:
                    first_cond = conditions[0]
                    if first_cond in df_result.columns:
                        df_result.loc[mask, first_cond] = res['default_value']

    # --- Fill units in df_conditions ---
    for idx, row in df_conditions.iterrows():
        df_conditions.at[idx, 'condition_unit'] = row.get('condition_unit', "")

    return df_info, df_result, df_conditions

def apply_custom_values(df_info, df_conditions, df_result, customization_json):
    """
    Apply default/custom values from the customization JSON directly into the dataframes.
    This ensures that the Excel sheets are pre-filled.
    """
    # Parameters
    for param in customization_json.get("METADATA_PARAMETERS", []):
        mask = df_info["param_name"] == param.get("param_name")
        if mask.any() and "value" in param:
            df_info.loc[mask, "value"] = param["value"]

    # Sample preparations
    for prep in customization_json.get("METADATA_SAMPLE_PREP", []):
        mask = df_info["param_name"] == prep.get("param_sampleprep_name")
        if mask.any() and "value" in prep:
            df_info.loc[mask, "value"] = prep["value"]

    # Conditions
    for cond in customization_json.get("conditions", []):
        mask = df_conditions["conditon_name"] == cond.get("conditon_name")
        if mask.any() and "value" in cond:
            df_conditions.loc[mask, "value"] = cond["value"]

    # Results / Endpoints
    for result in customization_json.get("question3", []):
        mask = df_result["result_name"] == result.get("result_name")
        if mask.any() and "value" in result:
            df_result.loc[mask, "value"] = result["value"]

    return df_info, df_conditions, df_result


def _slug(text):
    """Mirror of data_entry_survey._slug — converts a param name to a question-name slug."""
    return _re.sub(r"[^A-Za-z0-9_]", "_", str(text).strip()).lower()


def write_customization_to_excel(file_path_xlsx, blueprint_json, custom_json):
    """
    Apply data-entry survey answers (custom_json) to an already-generated
    blueprint Excel file (file_path_xlsx).

    The custom_json is the SurveyJS answer payload produced by DataPage /
    data_entry_survey, whose question names follow the convention built by
    data_entry_survey.py::

        de_provenance_*          → provenance rows in Test_conditions
        de_sample__<slug>        → METADATA_SAMPLE_INFO rows
        de_sampleprep__<slug>    → METADATA_SAMPLE_PREP rows
        de_params__<slug>        → METADATA_PARAMETERS rows
        de_condition__<slug>     → condition level series (ignored for Excel
                                   column headers, used as data row values)
        de_results               → list-of-dicts, fills Results_TABLE rows
        de_raw_data              → list-of-dicts, fills Raw_data_TABLE rows
        de_calibration           → list-of-dicts, fills Calibration_TABLE rows
        de_experiment_id         → stored in hidden sheet
        de_notes                 → stored in hidden sheet

    Fills column B of Test_conditions for every matching param_name and
    appends data rows to Raw_data_TABLE / Results_TABLE / Calibration_TABLE.
    Also updates the hidden TemplateDesigner sheet so NeXus conversion picks
    up the merged data automatically.

    Parameters
    ----------
    file_path_xlsx : str
        Path to the Excel file previously generated by iom_format_2excel.
    blueprint_json : dict
        The original blueprint JSON (used to build slug→param_name maps).
    custom_json : dict
        The SurveyJS answer payload from the data-entry survey.

    Returns
    -------
    str
        Path to the modified Excel file.
    """
    from openpyxl import load_workbook
    import json
    # data_only=False otherwise we don't see the forulae
    wb = load_workbook(file_path_xlsx, data_only=False)

    # ------------------------------------------------------------------
    # 1.  Build reverse-slug maps:  slug → original param_name
    # ------------------------------------------------------------------
    slug_to_param   = {}   # de_params__<slug>    → param_name
    slug_to_prep    = {}   # de_sampleprep__<slug> → param_sampleprep_name
    slug_to_sample  = {}   # de_sample__<slug>    → param_sample_name

    for p in blueprint_json.get(METADATA_PARAMETERS, []):
        name = p.get("param_name", "")
        slug_to_param[f"de_params__{_slug(name)}"] = name

    for p in blueprint_json.get(METADATA_SAMPLE_PREP, []):
        name = p.get("param_sampleprep_name", "")
        slug_to_prep[f"de_sampleprep__{_slug(name)}"] = name

    for p in blueprint_json.get(METADATA_SAMPLE_INFO, []):
        name = p.get("param_sample_name", "")
        slug_to_sample[f"de_sample__{_slug(name)}"] = name

    # Provenance keys map directly to method-metadata row labels
    # (these are the labels written by get_method_metadata / iom_format_2excel)
    PROVENANCE_MAP = {
        "de_provenance_operator":     "Assay/Test work conducted by",
        "de_provenance_provider":     "Partner conducting test/assay",
        "de_provenance_contact":      "Lead Scientist & contact for test",
        "de_provenance_project":      "Project Work Package",
        "de_provenance_workpackage":  "Project Work Package",
        "de_provenance_startdate":    "Test start date",
        "de_provenance_enddate":      "Test end date",
    }

    # ------------------------------------------------------------------
    # 2.  Fill Test_conditions column B
    # ------------------------------------------------------------------
    _SHEET_CONDITIONS = "Test_conditions"
    if _SHEET_CONDITIONS in wb.sheetnames:
        ws_tc = wb[_SHEET_CONDITIONS]

        # Build row-index map:  param_name (col A text) → row number
        label_to_row = {}
        for row in ws_tc.iter_rows():
            a_cell = row[0]
            if a_cell.value and str(a_cell.value).strip():
                label_to_row[str(a_cell.value).strip()] = a_cell.row

        def _write_param(label, value):
            """Write value into column B of the row whose column-A == label."""
            if value is None or value == "":
                return
            row_num = label_to_row.get(label)
            if row_num is not None:
                cell = ws_tc.cell(row=row_num, column=2)
                # do not overwrite formulas
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    print(cell.value)
                    return
                cell.value = str(value)

        for q_name, value in custom_json.items():
            # Provenance
            if q_name in PROVENANCE_MAP:
                _write_param(PROVENANCE_MAP[q_name], value)
            # Method parameters
            elif q_name in slug_to_param:
                _write_param(slug_to_param[q_name], value)
            # Sample prep parameters
            elif q_name in slug_to_prep:
                _write_param(slug_to_prep[q_name], value)
            # Sample info parameters
            elif q_name in slug_to_sample:
                _write_param(slug_to_sample[q_name], value)

    # ------------------------------------------------------------------
    # 3.  Build slug→endpoint name maps for data tables
    # ------------------------------------------------------------------
    def _ep_slug_map(items, name_key):
        """Return dict: ep__<slug> → endpoint_name."""
        return {f"ep__{_slug(item.get(name_key,''))}": item.get(name_key, "")
                for item in items}

    def _cond_slug_map(conditions):
        """Return dict: cond__<slug> → condition_name."""
        return {f"cond__{_slug(c.get('conditon_name',''))}": c.get("conditon_name", "")
                for c in conditions}

    ep_results_map  = _ep_slug_map(blueprint_json.get("question3", []),      "result_name")
    ep_raw_map      = _ep_slug_map(blueprint_json.get("raw_data_report", []),"raw_endpoint")
    ep_cal_map      = _ep_slug_map(blueprint_json.get("calibration_report", []), "calibration_entry")
    cond_map        = _cond_slug_map(blueprint_json.get("conditions", []))

    # ------------------------------------------------------------------
    # 4.  Append data rows to table sheets
    # ------------------------------------------------------------------
    def _find_header_row(ws):
        """Return the last contiguous header row index (1-based) before data."""
        # Tables have 2 header rows (name, unit) then data from row 3
        return 2

    def _build_col_map(ws, ep_map, cond_map):
        """
        Map column letter → original name for a MultiIndex-header sheet.
        Row 1 = top-level name, row 2 = unit (may start with 'Unnamed').
        Returns dict: original_name → col_index (1-based)
        """
        col_map = {}   # original_name → openpyxl column index (1-based)
        for cell in ws[1]:   # row 1 = endpoint / condition names
            if cell.value and not str(cell.value).startswith("Unnamed"):
                col_map[str(cell.value).strip()] = cell.column
        return col_map

    def _append_rows(ws, rows_data, ep_map, cond_map):
        """
        Append survey matrix rows to an openpyxl worksheet.

        rows_data : list[dict] — each dict is one survey matrixdynamic row.
          keys follow the pattern  cond__<slug>  or  ep__<slug>  or ep__<slug>__err
        ep_map   : {ep__<slug>: endpoint_name}
        cond_map : {cond__<slug>: condition_name}
        """
        if not rows_data:
            return

        col_map = _build_col_map(ws, ep_map, cond_map)
        next_row = ws.max_row + 1

        for row_dict in rows_data:
            for q_key, value in row_dict.items():
                # Resolve slug key to original column name
                original_name = None
                if q_key in ep_map:
                    original_name = ep_map[q_key]
                elif q_key in cond_map:
                    original_name = cond_map[q_key]
                elif q_key.endswith("__err"):
                    # Uncertainty column: ep__<slug>__err → endpoint_name ± ...
                    # The header contains the endpoint name; match by prefix
                    base_slug = q_key[:-5]   # strip __err
                    base_name = ep_map.get(base_slug, "")
                    # Find the uncertainty column (header contains base_name and err/SD)
                    for col_name, col_idx in col_map.items():
                        if base_name in col_name and col_name != base_name:
                            original_name = col_name
                            break

                if original_name and original_name in col_map:
                    col_idx = col_map[original_name]
                    ws.cell(row=next_row, column=col_idx, value=value)

            next_row += 1

    # Results_TABLE
    if "de_results" in custom_json and "Results_TABLE" in wb.sheetnames:
        _append_rows(
            wb["Results_TABLE"],
            custom_json["de_results"],
            ep_results_map,
            cond_map
        )

    # Raw_data_TABLE
    if "de_raw_data" in custom_json and "Raw_data_TABLE" in wb.sheetnames:
        _append_rows(
            wb["Raw_data_TABLE"],
            custom_json["de_raw_data"],
            ep_raw_map,
            cond_map
        )

    # Calibration_TABLE
    if "de_calibration" in custom_json and "Calibration_TABLE" in wb.sheetnames:
        # Calibration has a standard_label column + endpoint columns
        ep_cal_with_label = dict(ep_cal_map)
        ep_cal_with_label["cal_standard_label"] = "Sample"
        _append_rows(
            wb["Calibration_TABLE"],
            custom_json["de_calibration"],
            ep_cal_with_label,
            {}
        )

    # ------------------------------------------------------------------
    # 5.  Update hidden TemplateDesigner sheet with merged blueprint+custom
    # ------------------------------------------------------------------
    if "TemplateDesigner" in wb.sheetnames:
        ws_td = wb["TemplateDesigner"]
        ws_td["A4"] = "customization"
        ws_td["B4"] = json.dumps(custom_json)
        ws_td["A5"] = "experiment_id"
        ws_td["B5"] = custom_json.get("de_experiment_id", "")
        ws_td["A6"] = "notes"
        ws_td["B6"] = custom_json.get("de_notes", "")

    # ------------------------------------------------------------------
    # 6.  Duplicate material columns based on de_material_count
    # ------------------------------------------------------------------
    n_materials = int(custom_json.get("de_material_count", 1) or 1)
    if n_materials > 1 and "Test_conditions" in wb.sheetnames:
        _duplicate_material_columns(wb["Test_conditions"], n_materials)

    # ------------------------------------------------------------------
    # 7.  Add optional example material row to Materials sheet
    # ------------------------------------------------------------------
    _add_example_material_to_sheet(
        wb,
        blueprint_json.get(METADATA_SAMPLE_INFO, []),
        custom_json
    )

    wb.save(file_path_xlsx)
    return file_path_xlsx


def _add_example_material_to_sheet(wb, sample_info_params, custom_json):
    """
    If the user filled in the optional example material fields, write one
    example data row to the Materials sheet without changing column headers.
    """
    if "Materials" not in wb.sheetnames:
        return

    ws_mat = wb["Materials"]
    example = {}
    for param in sample_info_params:
        name = param.get("param_sample_name", "")
        group = param.get("param_sample_group", "")
        slug_key = f"de_sample__{_slug(name)}"
        value = custom_json.get(slug_key, "")
        if value:
            example[group] = value

    if not example:
        return

    # Column positions matching get_materials_columns() (1-based):
    # ["", "ERM identifier", "ID", "Name", "CAS", "type",
    #  "Supplier", "Supplier code", "Batch", "Core", "BET surface in m²/g"]
    GROUP_TO_COL = {
        "ID":       3,
        "NAME":     4,
        "CASRN":    5,
        "SUPPLIER": 7,
        "BATCH":    9,
    }
    next_row = ws_mat.max_row + 1
    for group, value in example.items():
        col = GROUP_TO_COL.get(group)
        if col:
            ws_mat.cell(row=next_row, column=col, value=value)


from openpyxl.formula.translate import Translator
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def _get_cell_formula(cell):
    """Return the formula string from a cell if present, otherwise None."""
    val = cell.value
    if isinstance(val, str) and val.startswith("="):
        return val
    raw = getattr(cell, "_value", None)
    if isinstance(raw, str) and raw.startswith("="):
        return raw
    return None


def _duplicate_material_columns(ws_tc, n_materials):
    """
    Duplicate the value column (next to 'Test Material Details' labels)
    so that n_materials columns are present side-by-side.
    Preserves and correctly shifts relative formulas and Data Validations.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.formula.translate import Translator
    from copy import copy

    if n_materials <= 1:
        return

    block_start = None
    block_rows = []
    in_block = False

    # --- Locate the "Test Material Details" block ---
    # We iterate and collect rows until we hit the NEXT explicit section header
    for row in ws_tc.iter_rows():
        a_val = row[0].value
        a_text = str(a_val).strip() if a_val else ""

        if a_text == "Test Material Details":
            block_start = row[0].row
            in_block = True
            continue

        if in_block:
            # If we find a new section header (bold/filled cell in Col A) 
            # and we already have some rows, we stop.
            # Adjust the condition below if your headers have a specific pattern.
            if a_text and any(header in a_text for header in ["Results", "Raw Data", "Section"]):
                break
            
            # We include the row if there is a label OR if there is a value/formula in Col B
            # This ensures we don't skip rows where Column A is empty but Column B has a formula
            b_val = row[1].value
            if a_text or b_val:
                block_rows.append(row[0].row)

    if not block_rows:
        return

    SOURCE_VALUE_COL = 2  # column B
    source_letter = get_column_letter(SOURCE_VALUE_COL)

    # --- Duplicate columns ---
    for mat_idx in range(1, n_materials):
        target_col = SOURCE_VALUE_COL + mat_idx
        target_letter = get_column_letter(target_col)

        # 1. Update Header Row
        if block_start is not None:
            ws_tc.cell(
                row=block_start,
                column=target_col,
                value=f"Test Material Details ({mat_idx + 1})"
            )

        # 2. Copy Values, Formulas, and Styles
        for row_num in block_rows:
            src_cell = ws_tc.cell(row=row_num, column=SOURCE_VALUE_COL)
            dst_cell = ws_tc.cell(row=row_num, column=target_col)

            # Handle Formulas
            formula = _get_cell_formula(src_cell)
            if formula:
                dst_cell.value = Translator(
                    formula,
                    origin=src_cell.coordinate
                ).translate_formula(dst_cell.coordinate)
            else:
                dst_cell.value = src_cell.value
            
            # Essential: Copy styles (dropdown arrows and borders)
            if src_cell.has_style:
                dst_cell.style = copy(src_cell.style)

        # 3. Duplicate Data Validations
        for dv in ws_tc.data_validations.dataValidation:
            for row_num in block_rows:
                src_coord = f"{source_letter}{row_num}"
                if src_coord in dv.cells:
                    dv.add(f"{target_letter}{row_num}")