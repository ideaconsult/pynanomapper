import uuid
import pyambit.datamodel as mx
import pandas as pd
import numbers
import re
from typing import IO, List
from openpyxl.utils import get_column_letter
import numpy as np
from pynanomapper.datamodel.templates.template_config import TemplateDesignerConfig


class TemplateDesignerParser(TemplateDesignerConfig):
    """Parser to convert TemplateDesigner Excel files into AMBIT data model objects."""

    def __init__(self, xlsx_file: IO):
        self.template_json = self.parse_hidden(xlsx_file)
        tc = pd.read_excel(xlsx_file, sheet_name="Test_conditions", header=None)
        tc.columns = [get_column_letter(i+1) for i in range(tc.shape[1])]
        self.test_conditions = tc
        self.materials = pd.read_excel(xlsx_file, sheet_name="Materials") 
        _data_sheets = self.template_json["data_sheets"]    
        if "data_raw" in _data_sheets:
            self.raw = pd.read_excel(xlsx_file, sheet_name="Raw_data_TABLE", header=[0,1]) 
        else: 
            self.raw = None
        if "data_processed" in _data_sheets:
            self.results = pd.read_excel(xlsx_file, sheet_name="Results_TABLE", header=[0,1])
        else:
            self.results = None
        if "data_calibration" in _data_sheets:
            self.calibration = pd.read_excel(xlsx_file, sheet_name="Calibration_TABLE", header=[0,1])
        else:
            self.calibration = None

   
    def parse_value_unit(self, s, unit=None):
        """
        Parses a string with a numeric value followed by a unit.
        
        Args:
            s: Input string, e.g. "12.5 mg", "100kg", "3.2e-4 mol"
            
        Returns:
            tuple: (value as float, unit as string)
                Returns (s, "") if input is not a string.
        """
        if not isinstance(s, str):
            # If already a number or None, return as value with empty unit
            return s, unit
        
        s = s.strip()
        match = re.match(r"([-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)\s*(.*)", s)
        if match:
            value = float(match.group(1))
            unit = match.group(2).strip()
            return value, unit
        else:
            # Could not parse, return original string as unitless value
            return s, unit

    def _get_rows_from_match(self, df, search_text, n_rows=1, start_col="B"):
        """
        Finds the row in column 'A' matching search_text and returns a subset DataFrame
        starting from start_col to the first empty column, for n_rows downward.

        Args:
            df (pd.DataFrame): The DataFrame to search.
            search_text: Value to find in column 'A'.
            n_rows (int): Number of rows to take starting from the found row.
            start_col (str): Column to start from (default "B").

        Returns:
            pd.DataFrame: Subset of the DataFrame with header columns.
                        Empty DataFrame if search_text not found.
        """
        # Find the row index
        matching_rows = df[df["A"] == search_text]
        if matching_rows.empty:
            return pd.DataFrame(columns=df.columns)  # empty DF with same headers

        start_row_idx = matching_rows.index[0]
        end_row_idx = start_row_idx + n_rows

        # Determine the range of columns from start_col to first empty in the first matched row
        start_idx = df.columns.get_loc(start_col)
        row = df.iloc[start_row_idx]

        # Find first empty column after start_col
        end_idx = start_idx
        for col in df.columns[start_idx:]:
            if pd.isna(row[col]) or row[col] == "":
                break
            end_idx += 1

        subset_df = df.iloc[start_row_idx:end_row_idx, start_idx:end_idx]

        return subset_df

    def get_materials_used(self):
        return self._get_rows_from_match(self.test_conditions, "Select item from Project Materials list", n_rows=1)

    def get_protocol_application(self):
        # Define Protocol from template metadata
        protocol = mx.Protocol(
            topcategory=self.template_json["PROTOCOL_TOP_CATEGORY"],
            category=mx.EndpointCategory(code=self.template_json["PROTOCOL_CATEGORY_CODE"]),
            guideline=[self.template_json["METHOD"]]
        )
        pa = mx.ProtocolApplication(protocol=protocol, effects=[])
        pa.parameters = self.get_parameters()
        return pa

    def _add_excel_col_letters(self, c_df: pd.DataFrame, cols_array, col_letter="col_letter") -> pd.DataFrame:
        """
        Adds an Excel-style column letter to each row in c_df where the 'name' matches a column in cols_array.

        Parameters
        ----------
        c_df : pd.DataFrame
            DataFrame with a column 'name' containing the logical column names.
        cols_array : Index or list
            Column array (can be MultiIndex or list of strings).

        Returns
        -------
        pd.DataFrame
            c_df with an additional column 'col_letter' containing Excel letters.
        """
        # Build mapping from column names to Excel letters
        name_to_letter = {}
        for i, col in enumerate(cols_array):
            col_name = col[0] if isinstance(col, tuple) else col
            if col_name in c_df["name"].values:
                name_to_letter[col_name] = get_column_letter(i + 1)  # Excel is 1-based

        # Add 'col_letter' column
        c_df = c_df.copy()
        c_df[col_letter] = c_df["name"].map(name_to_letter)
        return c_df

    def get_condition_df(self):
        df = pd.DataFrame(self.template_json["conditions"])
        df = df.rename(columns=lambda x: x.replace("condition_", "", 1))
        df = df.rename(columns=lambda x: x.replace("conditon_", "", 1))
        if self.results is not None:
            df = self._add_excel_col_letters(df, self.results.columns, col_letter="results_pos")
        if self.raw is not None:            
            df = self._add_excel_col_letters(df, self.raw.columns, col_letter="raw_pos")
        return df

    def get_parameters(self):
        params = {}
        # 5. Add metadata parameters
        for tag in ["METADATA_PARAMETERS", "METADATA_SAMPLE_PREP"]:
            for p in self.template_json.get(tag, []):
                p_name = p.get("param_name", None)
                p_group = p.get("param_group", None)
                p_unit = p.get("param_unit", None)                
                if p_name is None:
                    continue
                value = self._get_rows_from_match(self.test_conditions, p_name, n_rows=1)
                if value.empty:
                    continue
                value = value.iloc[0,0]
                if p_unit is None:
                    if isinstance(value, numbers.Number):
                        params[f"{p_group}/{p_name}"] = mx.Value(loValue=value)
                    elif pd.isna(value):
                        pass
                    else:
                        params[f"{p_group}/{p_name}"] = value  # already string                    
                else:
                    _val, _unit = self.parse_value_unit(value, p_unit)
                    params[f"{p_group}/{p_name}"]  = mx.Value(loValue=_val, unit=_unit)
        return params

    # --- Detect actual columns in MultiIndex safely ---
    def pick_column(self, df, name):
        """Return first matching column tuple for level 0 == name; None if not found."""
        for c in df.columns:
            if isinstance(c, tuple) and c[0] == name:
                return c
            elif c == name:
                return c
        return None
    
    def df_to_nd_effectarray_multicol(
        self,
        df: pd.DataFrame,
        axes_dict: dict,
        main_signal: str,
        aux_signals: list = None,
        endpoint: str = None,
        endpointtype: str = None,
    ):
        """
        Build an nD EffectArray from a DataFrame with MultiIndex columns.

        Units are automatically taken from the second level of the MultiIndex.
        """

        # --- Get units from second level of MultiIndex ---
        def get_unit(col):
            if isinstance(col, tuple) and len(col) > 1:
                unit = col[1]
                if unit is None or str(unit).startswith("Unnamed"):
                    return None
                return unit
            return None

        axes_names = list(axes_dict.keys())
        # Filter columns that exist
        axis_cols = {ax: col for ax, col in ((ax, self.pick_column(df, ax)) for ax in axes_names) if col is not None}
        for ax in axes_names or []:
            col = self.pick_column(df, ax)
            if col is not None:
                axis_cols[ax] = col

        main_col = self.pick_column(df, main_signal)
        if main_col is None:
            raise ValueError(f"Main signal column '{main_signal}' not found in DataFrame")
        aux_cols = {}
        aux_signals = list(aux_signals) if aux_signals is not None else []
        for aux in aux_signals:
            col = self.pick_column(df, aux)
            if col is not None:
                aux_cols[aux] = col

        # --- Build axes ValueArrays dynamically ---
        axes_dict = {}
        for ax, col in axis_cols.items():
            values = pd.unique(df[col])
            _unit = get_unit(col)
            axes_dict[ax] = mx.ValueArray(values=values, unit=_unit)  # axes usually have their own units if needed

        # --- Prepare nD shape ---
        axes_list = list(axes_dict.keys())
        shape = tuple(len(axes_dict[ax].values) for ax in axes_list)

        # Build index maps
        idx_map = {ax: {val: i for i, val in enumerate(axes_dict[ax].values)} for ax in axes_list}

        # --- Initialize main signal and auxiliary matrices ---
        signal_matrix = np.full(shape, np.nan)
        aux_matrices = {}
        for aux_name, col_id in aux_cols.items():  # col_id is the MultiIndex tuple
            col_data = df[col_id]                 # get the actual column
            if np.issubdtype(col_data.dtype, np.number):
                aux_matrices[aux_name] = np.full(shape, np.nan)
            else:
                aux_matrices[aux_name] = np.full(shape, None, dtype=object)

        # --- Fill matrices ---
        for _, row in df.iterrows():
            idx = tuple(idx_map[ax][row[axis_cols[ax]]] for ax in axes_list)
            signal_matrix[idx] = row[main_col]
            for aux_name, aux_col in aux_cols.items():
                aux_matrices[aux_name][idx] = row[aux_col]

        main_unit = get_unit(main_col)
        auxiliary = {k: mx.ValueArray(values=v, unit=get_unit(aux_cols[k])) for k, v in aux_matrices.items()} if aux_matrices else None

        # --- Build main signal ValueArray ---
        signal_va = mx.ValueArray(
            values=signal_matrix,
            unit=main_unit,
            auxiliary=auxiliary
        )

        # --- Build EffectArray ---
        earray = mx.EffectArray(
            endpoint=endpoint or main_signal,
            endpointtype=endpointtype,
            conditions={},
            signal=signal_va,
            axes=axes_dict,
            axis_groups=None
        )
        return earray

    def parse(self) -> mx.Substances:
        _data_sheets = self.template_json["data_sheets"]    
        if "data_raw" in _data_sheets:
            self.parse_raw_data()
        if "data_processed" in _data_sheets:
            self.parse_processed_data()
        if "data_calibration" in _data_sheets:
            self.parse_calibration()

    def _parse_data(self, data=None, endpoints_df=None):
        if data is None:
            return
        for index, row in data.iterrows():
            for (name, unit) in data.columns:
                value = row[(name, unit)]   # <-- this gives the cell value
                _unit = None if unit.startswith("Unnamed") else unit
                if name == "Material":
                    print(value)
                else:
                    if pd.notna(value):
                        print(f"Row {index}, {name} [{_unit}] = {value}")
                                               
        return data, endpoints_df
    
    def parse_raw_data(self):
        if self.raw is None:
            return None, None
        else:
            df = self._get_endpoints_df_raw()
            return self._parse_data(self.raw, df)

    def parse_processed_data(self):
        if self.results is None:
            return None, None
        else:
            df = self._get_endpoints_df_results()
            return self._parse_data(self.results, df)

    def _get_endpoints_df(self, tag="raw_data_report"):
        df = pd.DataFrame(self.template_json[tag])
        df = df.rename(columns=lambda x: x.replace("raw_", "", 1))
        df = df.rename(columns={"endpoint": "name"})
        df = df.rename(columns=lambda x: x.replace("result_", "", 1))
        df = df.rename(columns=lambda x: x.replace("results_", "", 1))
        return df

    def get_endpoints_df_raw(self):
        df = self._get_endpoints_df(tag="raw_data_report")
        return self._add_excel_col_letters(df, self.raw.columns, col_letter="raw_pos")

    def get_endpoints_df_results(self):
        df = self._get_endpoints_df(tag="question3")
        return self._add_excel_col_letters(df, self.results.columns, col_letter="results_pos")

    def _get_config_effects(self, cols_array, conditions_df, endpoints,
                            col_pos="raw_pos", sheet_name=None):
        effects = []
        for i, col in enumerate(cols_array):
            e = endpoints.loc[endpoints["name"] == col[0]]
            if not e.empty:
                col_letter = get_column_letter(i + 1) 
                effectrecord = {"ENDPOINT" : col[0], 
                                "VALUE" : { "COLUMN_INDEX": col_letter}
                                }
                if sheet_name is not None:
                    effectrecord["SHEET_NAME"] = sheet_name
                endpoint_type = e.get("aggregate",None)
                if endpoint_type is not None:
                    if not pd.isna(endpoint_type.values[0]):
                        effectrecord["ENDPOINT_TYPE"] = endpoint_type.values[0]
                conditions = e.get("conditions",None)
                if conditions is not None:
                    effectrecord["CONDITIONS"] = {}
                    conditions = conditions.values[0]
                    for cond in conditions:
                        c = conditions_df.loc[conditions_df["name"] == cond]
                        if not c.empty:
                            effectrecord["CONDITIONS"][cond] = {"COLUMN_INDEX" : c[col_pos].values[0]}
                if not pd.isna(e["unit"].values[0] ):
                    effectrecord["UNIT"] = e["unit"].values[0]
                effects.append(effectrecord)
        return effects
    
    def get_config(self):
        if self.get_layout() == "dose_response":
            config = {}
            config["PROTOCOL_APPLICATIONS"] = []
            config["PROTOCOL_APPLICATIONS"].append(self.get_config_papp())
            return config
        else:
            raise Exception("Not implemented")

    def get_config_params(self):
        config = {}
        # 5. Add metadata parameters
        for tag in ["METADATA_PARAMETERS", "METADATA_SAMPLE_PREP"]:
            for p in self.template_json.get(tag, []):
                p_name = p.get("param_name",None)
                if p_name is None:
                    continue
                value = self._get_rows_from_match(self.test_conditions, p_name, n_rows=1)
                if value.empty:
                    continue
                p_group = p.get("param_group", None)
                p_unit = p.get("param_unit", None)            
                config[f"{p_group}/{p_name}"] = {"ITERATION": "ABSOLUTE_LOCATION",
                        "SHEET_INDEX": 1,
                        "COLUMN_INDEX": "B",
                        "ROW_INDEX": value.index[0]
                        }
                if p_unit is not None:
                    config[f"{p_group}/{p_name}"]["UNIT"] = p_unit
        return config

    def get_config_papp(self):
        return {
            "PROTOCOL_TOP_CATEGORY" : None,
			"PROTOCOL_ENDPOINT": None,
			"PROTOCOL_CATEGORY_CODE": None,
			"PROTOCOL_GUIDELINE": {
				"guideline1": {
					"COLUMN_INDEX": None
				}
			},
			"CITATION_YEAR": {
				"COLUMN_INDEX": None,
				"DATA_INTERPRETATION": "AS_DATE"
			},
			"CITATION_TITLE": {
				"COLUMN_INDEX": None
			},
			"CITATION_OWNER": {
				"COLUMN_INDEX": None
			},
            "PARAMETERS": self.get_config_params(),
            "EFFECTS": self.get_config_effects(),
        }

    def get_config_effects(self):
        effects_raw = []
        effects_result = []
        if self.raw is not None:
            effects_raw = self._get_config_effects(
                cols_array=self.raw.columns,
                conditions_df=self.get_condition_df(), 
                endpoints=self.get_endpoints_df_raw(),
                col_pos="raw_pos",
                sheet_name="Raw_data_TABLE")
        if self.results is not None:
            effects_result = self._get_config_effects(
                cols_array=self.results.columns,
                conditions_df=self.get_condition_df(), 
                endpoints=self.get_endpoints_df_results(), 
                col_pos="results_pos",
                sheet_name="Results_TABLE")
        effects_raw.extend(effects_result)
        return effects_raw

    def parse_calibration(self):
        return None

    def to_protocol_application(self) -> mx.ProtocolApplication:
        """
        Convert parsed Excel template to AMBIT ProtocolApplication.
        
        Returns:
            ProtocolApplication object with protocol, parameters, and effects
            
        Example:
            >>> parser = TemplateDesignerParser("template.xlsx")
            >>> pa = parser.to_protocol_application()
            >>> print(pa.protocol.category.code)
        """
        # Get protocol from template metadata
        protocol = mx.Protocol(
            topcategory=self.template_json.get("PROTOCOL_TOP_CATEGORY"),
            category=mx.EndpointCategory(
                code=self.template_json.get("PROTOCOL_CATEGORY_CODE", "")
            ),
            guideline=[self.template_json.get("EXPERIMENT", "")]
        )
        
        # Get parameters
        parameters = self.get_parameters()
        
        # Parse effects from raw and processed data
        effects = []
        
        if self.raw is not None:
            raw_effects = self._parse_effects_from_dataframe(
                df=self.raw,
                endpoints_df=self.get_endpoints_df_raw(),
                conditions_df=self.get_condition_df(),
                endpoint_type="RAW_DATA"
            )
            effects.extend(raw_effects)
        
        if self.results is not None:
            result_effects = self._parse_effects_from_dataframe(
                df=self.results,
                endpoints_df=self.get_endpoints_df_results(),
                conditions_df=self.get_condition_df(),
                endpoint_type="AGGREGATED"
            )
            effects.extend(result_effects)
        # Create ProtocolApplication
        pa = mx.ProtocolApplication(
            protocol=protocol,
            effects=effects,
            parameters=parameters,
            uuid=str(uuid.uuid4()),
            investigation_uuid=self.template_json.get("investigation_uuid", str(uuid.uuid4())),
            assay_uuid=self.template_json.get("assay_uuid", str(uuid.uuid4())),
            citation=mx.Citation(
                owner=self.template_json.get("provenance_provider", "Unknown"),
                title=self.template_json.get("EXPERIMENT", "Template Designer Export"),
                year=None
            )
        )
        
        return pa

    def to_substances(self) -> mx.Substances:
        """
        Convert parsed Excel to Substances with study data.
        
        Creates SubstanceRecord objects from the Materials sheet and attaches
        the protocol application as study data.
        
        Returns:
            Substances object containing all materials with study data
            
        Example:
            >>> parser = TemplateDesignerParser("template.xlsx")
            >>> substances = parser.to_substances()
            >>> print(f"Found {len(substances.substance)} substances")
        """
        substances = []
        
        # Get protocol application (shared across all materials)
        pa = self.to_protocol_application()
        
        # Parse materials from Materials sheet
        _materials_used = self.get_materials_used().iloc[0].to_numpy()
        #print(type(_materials_used), _materials_used)
        filtered_materials = self.materials[self.materials["ERM identifier"].isin(_materials_used)]

        if filtered_materials is not None and not filtered_materials.empty:
            for idx, material_row in filtered_materials.iterrows():
                # Extract material ID - try different column names
                material_id = None
                for id_col in ['ID', 'ERM identifier' ]:
                    if id_col in material_row and pd.notna(material_row[id_col]):
                        material_id = str(material_row[id_col])
                        break
                
                if material_id is None:
                    material_id = str(uuid.uuid4())
                
                # Extract material name
                material_name = None
                material_type = None
                for name_col in ['Name' ]:
                    if name_col in material_row and pd.notna(material_row[name_col]):
                        material_name = str(material_row[name_col])
                        material_type = str(material_row["type"])
                        break
                
                if material_name is None:
                    material_name = f"Material_{idx}"
                
                # Create SubstanceRecord
                substance = mx.SubstanceRecord(
                    i5uuid=material_id,
                    name=material_name,
                    publicname=material_name,
                    substanceType=material_type,  # Default, could be parameterized
                    ownerName=self.template_json.get("provenance_provider", "Unknown"),
                    ownerUUID=str(uuid.uuid4())
                )
                
                # Clone protocol application for this substance
                # Each substance gets its own PA with the same data
                pa_copy = mx.ProtocolApplication(
                    protocol=pa.protocol,
                    effects=pa.effects,
                    parameters=pa.parameters,
                    uuid=str(uuid.uuid4()),
                    investigation_uuid=pa.investigation_uuid,
                    assay_uuid=pa.assay_uuid,
                    citation=pa.citation,
                    owner=mx.SampleLink(
                        substance=mx.Sample(uuid=material_id),
                        company=mx.Company(name=substance.ownerName)
                    )
                )
                
                # Add protocol application as study
                substance.study = [pa_copy]
                substances.append(substance)
        else:
            raise Exception("No Materials sheet!")
        
        return mx.Substances(substance=substances)

    def _parse_effects_from_dataframe(
        self,
        df: pd.DataFrame,
        endpoints_df: pd.DataFrame,
        conditions_df: pd.DataFrame,
        endpoint_type: str = "RAW_DATA"
    ) -> List[mx.EffectArray]:
        """
        Parse effects from a data table (raw or results) into EffectArray objects.
        
        Args:
            df: DataFrame with multi-level column headers (endpoint, unit)
            endpoints_df: DataFrame with endpoint metadata
            conditions_df: DataFrame with condition metadata
            endpoint_type: Type of endpoint ("RAW_DATA" or "AGGREGATED")
            
        Returns:
            List of EffectArray objects
        """
        effects = []
        
        # Identify condition columns and endpoint columns
        condition_names = conditions_df["name"].tolist() if "name" in conditions_df else []
        
        # Get endpoint columns from endpoints_df
        endpoint_col = "name" if "name" in endpoints_df else "endpoint"
        if endpoint_col not in endpoints_df:
            return effects
        
        endpoint_names = endpoints_df[endpoint_col].tolist()
        
        # For each endpoint, create an EffectArray
        for endpoint_name in endpoint_names:
            # Find the column in df
            endpoint_col_tuple = self.pick_column(df, endpoint_name)
            if endpoint_col_tuple is None:
                continue
            
            # Get endpoint metadata
            endpoint_meta = endpoints_df[endpoints_df[endpoint_col] == endpoint_name]
            if endpoint_meta.empty:
                continue
            
            endpoint_meta = endpoint_meta.iloc[0]
            
            # Get unit
            unit = None
            if isinstance(endpoint_col_tuple, tuple) and len(endpoint_col_tuple) > 1:
                unit = endpoint_col_tuple[1] if not endpoint_col_tuple[1].startswith("Unnamed") else None
            
            # Get conditions for this endpoint
            endpoint_conditions = endpoint_meta.get("conditions", [])
            if endpoint_conditions is None:
                endpoint_conditions = []
            if isinstance(endpoint_conditions, float) and np.isnan(endpoint_conditions):
                endpoint_conditions = []
            axes_dict = {}
            for cond_name in endpoint_conditions:
                cond_col = self.pick_column(df, cond_name)
                if cond_col is not None:
                    # Get unique values for this condition
                    unique_vals = pd.unique(df[cond_col].dropna())
                    if len(unique_vals) > 0:
                        # Get unit for condition
                        cond_meta = conditions_df[conditions_df["name"] == cond_name]
                        cond_unit = None
                        if not cond_meta.empty:
                            cond_unit = cond_meta.iloc[0].get("unit", None)
                            cond_unit = None if pd.isna(cond_unit) else cond_unit

                        axes_dict[cond_name] = mx.ValueArray(
                            values=np.array(unique_vals),
                            unit=cond_unit
                        )
            
            # Build signal array
            if len(axes_dict) > 0:
                # Multi-dimensional data - use df_to_nd_effectarray_multicol
                try:
                    effect = self.df_to_nd_effectarray_multicol(
                        df=df,
                        axes_dict=axes_dict,
                        main_signal=endpoint_name,
                        endpoint=endpoint_name,
                        endpointtype=endpoint_type
                    )
                    effects.append(effect)
                except Exception as e:
                    print(f"Warning: Could not create EffectArray for {endpoint_name}: {e}")
            else:
                # Simple 1D data
                values = df[endpoint_col_tuple].dropna().values
                if len(values) > 0:
                    signal = mx.ValueArray(values=np.array(values), unit=unit)
                    effect = mx.EffectArray(
                        endpoint=endpoint_name,
                        endpointtype=endpoint_type,
                        signal=signal,
                        axes={},
                        conditions={}
                    )
                    effects.append(effect)
        
        return effects