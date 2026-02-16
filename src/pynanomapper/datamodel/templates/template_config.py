import pandas as pd
import re
from typing import IO
from openpyxl.utils import get_column_letter
import ast
import json
from pathlib import Path


class TemplateDesignerConfig:
    """Parser to convert TemplateDesigner Excel files into AMBIT data model objects."""

    def __init__(self, xlsx_file: Path):
        self.template_file_name = xlsx_file
        self.template_json = self.parse_hidden(xlsx_file)
        self.materials = pd.read_excel(xlsx_file, sheet_name="Materials")
        if self.get_layout() == "dose_response":
            tc = pd.read_excel(xlsx_file, sheet_name="Test_conditions", header=None)
            tc.columns = [get_column_letter(i+1) for i in range(tc.shape[1])]
            self.test_conditions = tc
            _data_sheets = self.template_json["data_sheets"]    
            if "data_raw" in _data_sheets:
                self.raw = pd.read_excel(xlsx_file, sheet_name="Raw_data_TABLE", header=[0,1]) 
            else:
                self.raw = None
            if "data_processed" in _data_sheets:
                self.results = pd.read_excel(xlsx_file, sheet_name="Results_TABLE", header=[0,1])
            else:
                self.results = None
            if "data_calibration" in _data_sheets:
                self.calibration = pd.read_excel(xlsx_file, sheet_name="Calibration_TABLE", header=[0,1])
            else:
                self.calibration = None
        else:
            tmp = pd.read_excel(xlsx_file, sheet_name="Results_TABLE", header=None)
            self.results = tmp.iloc[4:]
            new_header = [tmp.iloc[0, 0]] + [tmp.iloc[0, 1]] + tmp.iloc[1, 2:].tolist()
            self.results.columns = new_header            
            tmp = pd.read_excel(xlsx_file, sheet_name="Experimental_setup", header=None)
            self.test_conditions = tmp.iloc[4:]
            new_header = [tmp.iloc[0, 0]] + tmp.iloc[2, 1:].tolist()
            self.test_conditions.columns = new_header
            self.provider_info = pd.read_excel(xlsx_file, sheet_name="Provider_informations", header=None)

    def get_method(self):
        return self.template_json["METHOD"]
    
    def get_layout(self):
        return self.template_json["template_layout"]
    
    def parse_value_unit(self, s, unit=None):
        """
        Parses a string with a numeric value followed by a unit.
        
        Args:
            s: Input string, e.g. "12.5 mg", "100kg", "3.2e-4 mol"
            
        Returns:
            tuple: (value as float, unit as string)
                Returns (s, "") if input is not a string.
        """
        if not isinstance(s, str):
            # If already a number or None, return as value with empty unit
            return s, unit
        
        s = s.strip()
        match = re.match(r"([-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)\s*(.*)", s)
        if match:
            value = float(match.group(1))
            unit = match.group(2).strip()
            return value, unit
        else:
            # Could not parse, return original string as unitless value
            return s, unit
        
        
    def _get_rows_from_match(self, df, search_text, n_rows=1, start_col="B"):
        """
        Finds the row in column 'A' matching search_text and returns a subset DataFrame
        starting from start_col to the first empty column, for n_rows downward.

        Args:
            df (pd.DataFrame): The DataFrame to search.
            search_text: Value to find in column 'A'.
            n_rows (int): Number of rows to take starting from the found row.
            start_col (str): Column to start from (default "B").

        Returns:
            pd.DataFrame: Subset of the DataFrame with header columns.
                        Empty DataFrame if search_text not found.
        """
        # Find the row index
        matching_rows = df[df["A"] == search_text]
        if matching_rows.empty:
            return pd.DataFrame(columns=df.columns)  # empty DF with same headers

        start_row_idx = matching_rows.index[0]
        end_row_idx = start_row_idx + n_rows

        # Determine the range of columns from start_col to first empty in the first matched row
        start_idx = df.columns.get_loc(start_col)
        row = df.iloc[start_row_idx]

        # Find first empty column after start_col
        end_idx = start_idx
        for col in df.columns[start_idx:]:
            if pd.isna(row[col]) or row[col] == "":
                break
            end_idx += 1

        subset_df = df.iloc[start_row_idx:end_row_idx, start_idx:end_idx]

        return subset_df

    def get_materials_used(self):
        if self.get_layout() == "dose_response":
            return self._get_rows_from_match(self.test_conditions, "Select item from Project Materials list", n_rows=1)
        else:
            raise Exception('Not implemented')

    def _add_excel_col_letters(self, c_df: pd.DataFrame, cols_array, col_letter="col_letter") -> pd.DataFrame:
        """
        Adds an Excel-style column letter to each row in c_df where the 'name' matches a column in cols_array.

        Parameters
        ----------
        c_df : pd.DataFrame
            DataFrame with a column 'name' containing the logical column names.
        cols_array : Index or list
            Column array (can be MultiIndex or list of strings).

        Returns
        -------
        pd.DataFrame
            c_df with an additional column 'col_letter' containing Excel letters.
        """
        # Build mapping from column names to Excel letters
        name_to_letter = {}
        for i, col in enumerate(cols_array):
            col_name = col[0] if isinstance(col, tuple) else col
            if col_name in c_df["name"].values:
                name_to_letter[col_name] = get_column_letter(i + 1)  # Excel is 1-based

        # Add 'col_letter' column
        c_df = c_df.copy()
        c_df[col_letter] = c_df["name"].map(name_to_letter)
        return c_df

    def get_condition_df(self):
        conditions = self.template_json.get("conditions", None)
        if conditions is not None:
            df = pd.DataFrame(conditions)
            df = df.rename(columns=lambda x: x.replace("condition_", "", 1))
            df = df.rename(columns=lambda x: x.replace("conditon_", "", 1))
            if self.results is not None:
                df = self._add_excel_col_letters(df, self.results.columns, col_letter="results_pos")
            if self.raw is not None:            
                df = self._add_excel_col_letters(df, self.raw.columns, col_letter="raw_pos")
            return df
        else:
            return None

    # --- Detect actual columns in MultiIndex safely ---
    def pick_column(self, df, name):
        """Return first matching column tuple for level 0 == name; None if not found."""
        for c in df.columns:
            if isinstance(c, tuple) and c[0] == name:
                return c
            elif c == name:
                return c
        return None
    
    def _get_endpoints_df(self, tag="raw_data_report"):
        df = pd.DataFrame(self.template_json[tag])
        df = df.rename(columns=lambda x: x.replace("raw_", "", 1))
        df = df.rename(columns={"endpoint": "name"})
        df = df.rename(columns=lambda x: x.replace("result_", "", 1))
        df = df.rename(columns=lambda x: x.replace("results_", "", 1))
        return df

    def get_endpoints_df_raw(self):
        df = self._get_endpoints_df(tag="raw_data_report")
        return self._add_excel_col_letters(df, self.raw.columns, col_letter="raw_pos")

    def get_endpoints_df_results(self):
        df = self._get_endpoints_df(tag="question3")
        return self._add_excel_col_letters(df, self.results.columns, col_letter="results_pos")

    def _get_config_effects(self, cols_array, conditions_df, endpoints,
                            col_pos="raw_pos", sheet_name=None):
        effects = []
        for i, col in enumerate(cols_array):
            e = endpoints.loc[endpoints["name"] == col[0]]
            if not e.empty:
                col_letter = get_column_letter(i + 1) 
                effectrecord = {"ENDPOINT" : col[0], 
                                "VALUE" : { "COLUMN_INDEX": col_letter}
                                }
                if sheet_name is not None:
                    effectrecord["SHEET_NAME"] = sheet_name
                endpoint_type = e.get("aggregate",None)
                if endpoint_type is not None:
                    if not pd.isna(endpoint_type.values[0]):
                        effectrecord["ENDPOINT_TYPE"] = endpoint_type.values[0]
                conditions = e.get("conditions",None)
                if conditions is not None:
                    effectrecord["CONDITIONS"] = {}
                    conditions = conditions.values[0]
                    for cond in conditions:
                        c = conditions_df.loc[conditions_df["name"] == cond]
                        if not c.empty:
                            effectrecord["CONDITIONS"][cond] = {"COLUMN_INDEX" : c[col_pos].values[0]}
                if not pd.isna(e["unit"].values[0] ):
                    effectrecord["UNIT"] = e["unit"].values[0]
                effects.append(effectrecord)
        return effects
    
    def get_config(self):
        config = {
            "TEMPLATE_INFO": {
            "NAME": self.template_json["template_name"],
            "VERSION": self.template_json["template_date"],
            "TYPE": 1,
            "UUID" : self.template_json["template_uuid"],
            "template_layout": self.template_json["template_layout"]
            }
        }
        config["DATA_ACCESS"] = {
            "ITERATION": "ROW_SINGLE",
            "SHEET_INDEX": 2,
            "START_ROW": 3,
            "END_ROW_": 3,
            "START_HEADER_ROW": 1,
            "END_HEADER_ROW": 2,
            "ALLOW_EMPTY": True,
            "RECOGNITION": "BY_INDEX"
        }
        config["SUBSTANCE_RECORD"] = {
            "PUBLIC_NAME": {
                "COLUMN_INDEX": "A"
            },
            "SUBSTANCE_NAME": {
                "COLUMN_INDEX": "A"
            }
        }
        config["PROTOCOL_APPLICATIONS"] = []        
        if self.get_layout() == "dose_response":
            config["PROTOCOL_APPLICATIONS"].append(self.get_config_papp())
            return config
        else:
            raise Exception("Not implemented")        

    def get_config_params(self):
        config = {"E.Method" : {"ITERATION": "ABSOLUTE_LOCATION", "SHEET_INDEX": 1,
                                "COLUMN_INDEX": "B", "ROW_INDEX": 13}}
        # 5. Add metadata parameters
        for tag in ["METADATA_PARAMETERS", "METADATA_SAMPLE_PREP"]:
            for p in self.template_json.get(tag, []):
                p_name = p.get("param_name",None)
                if p_name is None:
                    continue
                value = self._get_rows_from_match(self.test_conditions, p_name, n_rows=1)
                if value.empty:
                    continue
                p_group = p.get("param_group", None)
                p_unit = p.get("param_unit", None)            
                config[f"{p_group}/{p_name}"] = {
                        "ITERATION": "ABSOLUTE_LOCATION",
                        "SHEET_INDEX": 1,
                        "COLUMN_INDEX": "B",
                        "ROW_INDEX": value.index[0]
                        }
                if p_unit is not None:
                    config[f"{p_group}/{p_name}"]["UNIT"] = p_unit
        return config

    def get_config_papp(self):
        code = self.template_json["PROTOCOL_CATEGORY_CODE"]
        top_code = self.template_json["PROTOCOL_TOP_CATEGORY"]
        return {
            "PROTOCOL_TOP_CATEGORY" : top_code,
			"PROTOCOL_ENDPOINT": {
                "ITERATION": "ABSOLUTE_LOCATION",
                "COLUMN_INDEX": "B",
                "SHEET_INDEX" : 1, "ROW_INDEX" : 13},
			"PROTOCOL_CATEGORY_CODE": code,
			"PROTOCOL_GUIDELINE": {
				"guideline1": {
                    "ITERATION": "ABSOLUTE_LOCATION",
                    "SHEET_INDEX": 1,
                    "COLUMN_INDEX": "B",
                    "ROW_INDEX": 20
				}
			},
			"CITATION_YEAR": {
                "ITERATION": "ABSOLUTE_LOCATION",
                "SHEET_INDEX": 1,
                "COLUMN_INDEX": "B",
                "ROW_INDEX": 22,
				"DATA_INTERPRETATION": "AS_DATE"
			},
			"CITATION_TITLE": {
                "ITERATION": "ABSOLUTE_LOCATION",
                "SHEET_INDEX": 1,
                "COLUMN_INDEX": "B",
                "ROW_INDEX": 8
			},
			"CITATION_OWNER": {
                "ITERATION": "ABSOLUTE_LOCATION",
                "SHEET_INDEX": 1,
                "COLUMN_INDEX": "B",
                "ROW_INDEX": 9
			},
            "PARAMETERS": self.get_config_params(),
            "EFFECTS": self.get_config_effects(),
        }

    def get_config_effects(self):
        effects_raw = []
        effects_result = []
        if self.raw is not None:
            effects_raw = self._get_config_effects(
                cols_array=self.raw.columns,
                conditions_df=self.get_condition_df(), 
                endpoints=self.get_endpoints_df_raw(),
                col_pos="raw_pos",
                sheet_name="Raw_data_TABLE")
        if self.results is not None:
            effects_result = self._get_config_effects(
                cols_array=self.results.columns,
                conditions_df=self.get_condition_df(), 
                endpoints=self.get_endpoints_df_results(), 
                col_pos="results_pos",
                sheet_name="Results_TABLE")
        effects_raw.extend(effects_result)
        return effects_raw
        
    def parse_hidden(self, xlsx_file: IO):
        template_df = pd.read_excel(xlsx_file, sheet_name="TemplateDesigner")
        _json = template_df.iloc[0]["surveyjs"]
        try:
            return json.loads(_json.replace("'", "\""))
        except Exception as err:
            # Handle legacy files with single quotes instead of double quotes
            print(f"JSON parsing failed, trying ast.literal_eval: {err}")
            data = ast.literal_eval(_json)
            good_json_str = json.dumps(data, ensure_ascii=False, indent=2)
            return json.loads(good_json_str)

    def parse_calibration(self):
        return None